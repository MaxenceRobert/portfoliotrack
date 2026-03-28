import os
import csv
import io
import json
import datetime
from config import Config

def get_db():
    database_url = os.environ.get('DATABASE_URL')
    if database_url:
        import psycopg2
        import psycopg2.extras
        conn = psycopg2.connect(database_url)
        conn.autocommit = False
        return conn
    else:
        import sqlite3
        conn = sqlite3.connect(Config.DATABASE)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        return conn

def is_postgres():
    return bool(os.environ.get('DATABASE_URL'))

def placeholder():
    return '%s' if is_postgres() else '?'

def init_db():
    conn = get_db()
    c = conn.cursor()
    pg = is_postgres()

    if pg:
        c.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id            SERIAL PRIMARY KEY,
                email         TEXT    NOT NULL UNIQUE,
                password_hash TEXT    NOT NULL,
                created_at    TIMESTAMP DEFAULT NOW()
            )
        ''')
        c.execute('''
            ALTER TABLE users ADD COLUMN IF NOT EXISTS email_verified BOOLEAN DEFAULT FALSE
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS assets (
                id          SERIAL PRIMARY KEY,
                user_id     INTEGER NOT NULL,
                isin        TEXT,
                ticker      TEXT    NOT NULL,
                name        TEXT    NOT NULL,
                asset_type  TEXT    NOT NULL DEFAULT 'ETF',
                currency    TEXT    NOT NULL DEFAULT 'EUR',
                workspace   TEXT    DEFAULT 'perso',
                taux_fixe   REAL    DEFAULT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
                UNIQUE(user_id, ticker)
            )
        ''')
        # Migration : ajout colonne taux_fixe si absente
        try:
            c.execute('ALTER TABLE assets ADD COLUMN IF NOT EXISTS taux_fixe REAL DEFAULT NULL')
        except Exception:
            pass
        c.execute('''
            CREATE TABLE IF NOT EXISTS alternative_assets (
                id                SERIAL PRIMARY KEY,
                user_id           INTEGER NOT NULL,
                name              TEXT NOT NULL,
                category          TEXT NOT NULL DEFAULT 'Autre',
                acquisition_value REAL NOT NULL DEFAULT 0,
                current_value     REAL NOT NULL DEFAULT 0,
                acquisition_date  TEXT,
                notes             TEXT DEFAULT '',
                workspace         TEXT DEFAULT 'perso',
                created_at        TIMESTAMP DEFAULT NOW(),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS asset_catalog (
                ticker TEXT PRIMARY KEY, name TEXT NOT NULL,
                category TEXT, subcategory TEXT, region TEXT,
                description TEXT, currency TEXT DEFAULT 'EUR')
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS purchases (
                id              SERIAL PRIMARY KEY,
                user_id         INTEGER NOT NULL,
                asset_id        INTEGER NOT NULL,
                date            TEXT    NOT NULL,
                shares          REAL    NOT NULL,
                price_per_share REAL    NOT NULL,
                total_cost      REAL    NOT NULL,
                fees            REAL    DEFAULT 0,
                notes           TEXT    DEFAULT '',
                created_at      TIMESTAMP DEFAULT NOW(),
                FOREIGN KEY (user_id)  REFERENCES users(id)   ON DELETE CASCADE,
                FOREIGN KEY (asset_id) REFERENCES assets(id)  ON DELETE CASCADE
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS sales (
                id                SERIAL PRIMARY KEY,
                user_id           INTEGER NOT NULL,
                asset_id          INTEGER NOT NULL,
                date              TEXT    NOT NULL,
                shares            REAL    NOT NULL,
                price_per_share   REAL    NOT NULL,
                total_proceeds    REAL    NOT NULL,
                fees              REAL    DEFAULT 0,
                notes             TEXT    DEFAULT '',
                created_at        TIMESTAMP DEFAULT NOW(),
                FOREIGN KEY (user_id)  REFERENCES users(id)  ON DELETE CASCADE,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS dca_goals (
                id             SERIAL PRIMARY KEY,
                user_id        INTEGER NOT NULL UNIQUE,
                monthly_target REAL    NOT NULL DEFAULT 0,
                updated_at     TIMESTAMP DEFAULT NOW(),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS dividends (
                id          SERIAL PRIMARY KEY,
                user_id     INTEGER NOT NULL,
                asset_id    INTEGER NOT NULL,
                date        TEXT    NOT NULL,
                amount      REAL    NOT NULL,
                notes       TEXT    DEFAULT '',
                created_at  TIMESTAMP DEFAULT NOW(),
                FOREIGN KEY (user_id)  REFERENCES users(id)  ON DELETE CASCADE,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS email_tokens (
                id         SERIAL PRIMARY KEY,
                user_id    INTEGER NOT NULL,
                token      TEXT    NOT NULL UNIQUE,
                created_at TIMESTAMP DEFAULT NOW(),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS profil_investisseur (
                id             SERIAL PRIMARY KEY,
                user_id        INTEGER NOT NULL,
                score_global   INTEGER NOT NULL,
                nom_profil     TEXT    NOT NULL,
                scores_axes    TEXT    NOT NULL DEFAULT '{}',
                recommandation TEXT    NOT NULL DEFAULT '',
                date_test      TIMESTAMP DEFAULT NOW(),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS asset_risk_scores (
                id          SERIAL PRIMARY KEY,
                ticker      TEXT NOT NULL UNIQUE,
                score       INTEGER NOT NULL,
                volatilite  REAL,
                drawdown    REAL,
                beta        REAL,
                date_calcul TIMESTAMP DEFAULT NOW(),
                source      TEXT NOT NULL DEFAULT 'default'
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS reset_tokens (
                id         SERIAL PRIMARY KEY,
                user_id    INTEGER NOT NULL,
                token      TEXT    NOT NULL UNIQUE,
                expires_at TIMESTAMP NOT NULL,
                used       BOOLEAN DEFAULT FALSE,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS auto_dividends_cache (
                id        SERIAL PRIMARY KEY,
                ticker    TEXT NOT NULL UNIQUE,
                data_json TEXT NOT NULL,
                cached_at TIMESTAMP DEFAULT NOW()
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS envelopes (
                id             SERIAL PRIMARY KEY,
                user_id        INTEGER NOT NULL,
                type           TEXT    NOT NULL,
                nom            TEXT    NOT NULL,
                solde          REAL    NOT NULL DEFAULT 0,
                taux_annuel    REAL    NOT NULL DEFAULT 0,
                plafond        REAL,
                date_ouverture TEXT,
                created_at     TIMESTAMP DEFAULT NOW(),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        ''')
    else:
        c.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                email         TEXT    NOT NULL UNIQUE,
                password_hash TEXT    NOT NULL,
                created_at    TEXT    DEFAULT (datetime('now'))
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS assets (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id     INTEGER NOT NULL,
                isin        TEXT,
                ticker      TEXT,
                name        TEXT    NOT NULL,
                asset_type  TEXT    NOT NULL DEFAULT 'ETF',
                currency    TEXT    NOT NULL DEFAULT 'EUR',
                workspace   TEXT    DEFAULT 'perso',
                taux_fixe   REAL    DEFAULT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        ''')
        # Migration SQLite : ajout colonne taux_fixe si absente
        try:
            c.execute('ALTER TABLE assets ADD COLUMN taux_fixe REAL DEFAULT NULL')
        except Exception:
            pass
        c.execute('''
            CREATE TABLE IF NOT EXISTS alternative_assets (
                id                INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id           INTEGER NOT NULL,
                name              TEXT NOT NULL,
                category          TEXT NOT NULL DEFAULT 'Autre',
                acquisition_value REAL NOT NULL DEFAULT 0,
                current_value     REAL NOT NULL DEFAULT 0,
                acquisition_date  TEXT,
                notes             TEXT DEFAULT '',
                workspace         TEXT DEFAULT 'perso',
                created_at        TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS asset_catalog (
                ticker TEXT PRIMARY KEY, name TEXT NOT NULL,
                category TEXT, subcategory TEXT, region TEXT,
                description TEXT, currency TEXT DEFAULT 'EUR')
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS purchases (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id         INTEGER NOT NULL,
                asset_id        INTEGER NOT NULL,
                date            TEXT    NOT NULL,
                shares          REAL    NOT NULL,
                price_per_share REAL    NOT NULL,
                total_cost      REAL    NOT NULL,
                fees            REAL    DEFAULT 0,
                notes           TEXT    DEFAULT '',
                created_at      TEXT    DEFAULT (datetime('now')),
                FOREIGN KEY (user_id)  REFERENCES users(id)   ON DELETE CASCADE,
                FOREIGN KEY (asset_id) REFERENCES assets(id)  ON DELETE CASCADE
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS sales (
                id                INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id           INTEGER NOT NULL,
                asset_id          INTEGER NOT NULL,
                date              TEXT    NOT NULL,
                shares            REAL    NOT NULL,
                price_per_share   REAL    NOT NULL,
                total_proceeds    REAL    NOT NULL,
                fees              REAL    DEFAULT 0,
                notes             TEXT    DEFAULT '',
                created_at        TEXT    DEFAULT (datetime('now')),
                FOREIGN KEY (user_id)  REFERENCES users(id)  ON DELETE CASCADE,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS dca_goals (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id        INTEGER NOT NULL UNIQUE,
                monthly_target REAL    NOT NULL DEFAULT 0,
                updated_at     TEXT    DEFAULT (datetime('now')),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS dividends (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id     INTEGER NOT NULL,
                asset_id    INTEGER NOT NULL,
                date        TEXT    NOT NULL,
                amount      REAL    NOT NULL,
                notes       TEXT    DEFAULT '',
                created_at  TEXT    DEFAULT (datetime('now')),
                FOREIGN KEY (user_id)  REFERENCES users(id)  ON DELETE CASCADE,
                FOREIGN KEY (asset_id) REFERENCES assets(id) ON DELETE CASCADE
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS email_tokens (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id    INTEGER NOT NULL,
                token      TEXT    NOT NULL UNIQUE,
                created_at TEXT    DEFAULT (datetime('now')),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS profil_investisseur (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id        INTEGER NOT NULL,
                score_global   INTEGER NOT NULL,
                nom_profil     TEXT    NOT NULL,
                scores_axes    TEXT    NOT NULL DEFAULT '{}',
                recommandation TEXT    NOT NULL DEFAULT '',
                date_test      TEXT    DEFAULT (datetime('now')),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS asset_risk_scores (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                ticker      TEXT NOT NULL UNIQUE,
                score       INTEGER NOT NULL,
                volatilite  REAL,
                drawdown    REAL,
                beta        REAL,
                date_calcul TEXT DEFAULT (datetime('now')),
                source      TEXT NOT NULL DEFAULT 'default'
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS reset_tokens (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id    INTEGER NOT NULL,
                token      TEXT    NOT NULL UNIQUE,
                expires_at TEXT    NOT NULL,
                used       INTEGER DEFAULT 0,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS auto_dividends_cache (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                ticker    TEXT NOT NULL UNIQUE,
                data_json TEXT NOT NULL,
                cached_at TEXT DEFAULT (datetime('now'))
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS envelopes (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id        INTEGER NOT NULL,
                type           TEXT    NOT NULL,
                nom            TEXT    NOT NULL,
                solde          REAL    NOT NULL DEFAULT 0,
                taux_annuel    REAL    NOT NULL DEFAULT 0,
                plafond        REAL,
                date_ouverture TEXT,
                created_at     TEXT    DEFAULT (datetime('now')),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        ''')

    conn.commit()

    # ── Migration : ajout des colonnes sharpe + var_95, vidage cache si nouveau ──
    new_cols_added = False
    for col in ('sharpe', 'var_95'):
        try:
            if is_postgres():
                c.execute(f'ALTER TABLE asset_risk_scores ADD COLUMN IF NOT EXISTS {col} REAL')
            else:
                c.execute(f'ALTER TABLE asset_risk_scores ADD COLUMN {col} REAL')
            conn.commit()
            new_cols_added = True
            print(f"[migration] asset_risk_scores: colonne '{col}' ajoutée")
        except Exception:
            conn.rollback()

    if new_cols_added:
        c.execute('DELETE FROM asset_risk_scores')
        conn.commit()
        print("[migration] asset_risk_scores: cache vidé (nouvel algorithme de scoring)")

    # ── Migration : colonne envelope sur assets ───────────────────────────────
    try:
        if is_postgres():
            c.execute('ALTER TABLE assets ADD COLUMN IF NOT EXISTS envelope TEXT')
        else:
            c.execute('ALTER TABLE assets ADD COLUMN envelope TEXT')
        conn.commit()
        print("[migration] assets: colonne 'envelope' ajoutée")
    except Exception:
        conn.rollback()

    # ── Migration : colonne onboarding_completed sur users ───────────────────
    try:
        if is_postgres():
            c.execute('ALTER TABLE users ADD COLUMN IF NOT EXISTS onboarding_completed BOOLEAN DEFAULT FALSE')
        else:
            c.execute('ALTER TABLE users ADD COLUMN onboarding_completed INTEGER DEFAULT 0')
        conn.commit()
        print("[migration] users: colonne 'onboarding_completed' ajoutée")
    except Exception:
        conn.rollback()

    # ── Migration : table reset_tokens ───────────────────────────────────────
    try:
        if is_postgres():
            c.execute('''
                CREATE TABLE IF NOT EXISTS reset_tokens (
                    id         SERIAL PRIMARY KEY,
                    user_id    INTEGER NOT NULL,
                    token      TEXT    NOT NULL UNIQUE,
                    expires_at TIMESTAMP NOT NULL,
                    used       BOOLEAN DEFAULT FALSE,
                    FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                )
            ''')
        else:
            c.execute('''
                CREATE TABLE IF NOT EXISTS reset_tokens (
                    id         INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id    INTEGER NOT NULL,
                    token      TEXT    NOT NULL UNIQUE,
                    expires_at TEXT    NOT NULL,
                    used       INTEGER DEFAULT 0,
                    FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                )
            ''')
        conn.commit()
    except Exception:
        conn.rollback()

    # ── Migration : table auto_dividends_cache ────────────────────────────────
    try:
        if is_postgres():
            c.execute('''
                CREATE TABLE IF NOT EXISTS auto_dividends_cache (
                    id        SERIAL PRIMARY KEY,
                    ticker    TEXT NOT NULL UNIQUE,
                    data_json TEXT NOT NULL,
                    cached_at TIMESTAMP DEFAULT NOW()
                )
            ''')
        else:
            c.execute('''
                CREATE TABLE IF NOT EXISTS auto_dividends_cache (
                    id        INTEGER PRIMARY KEY AUTOINCREMENT,
                    ticker    TEXT NOT NULL UNIQUE,
                    data_json TEXT NOT NULL,
                    cached_at TEXT DEFAULT (datetime('now'))
                )
            ''')
        conn.commit()
    except Exception:
        conn.rollback()

    # ── Migration : colonne workspace sur assets ──────────────────────────────
    try:
        if is_postgres():
            c.execute('ALTER TABLE assets ADD COLUMN IF NOT EXISTS workspace TEXT DEFAULT \'perso\'')
        else:
            c.execute('ALTER TABLE assets ADD COLUMN workspace TEXT DEFAULT \'perso\'')
        conn.commit()
        print("[migration] assets: colonne 'workspace' ajoutée")
    except Exception:
        conn.rollback()

    # ── Migration : table alternative_assets ──────────────────────────────────
    try:
        if is_postgres():
            c.execute('''CREATE TABLE IF NOT EXISTS alternative_assets (
                id SERIAL PRIMARY KEY, user_id INTEGER NOT NULL,
                name TEXT NOT NULL, category TEXT NOT NULL DEFAULT 'Autre',
                acquisition_value REAL NOT NULL DEFAULT 0,
                current_value REAL NOT NULL DEFAULT 0,
                acquisition_date TEXT, notes TEXT DEFAULT '', workspace TEXT DEFAULT 'perso',
                created_at TIMESTAMP DEFAULT NOW(),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE)''')
        else:
            c.execute('''CREATE TABLE IF NOT EXISTS alternative_assets (
                id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL,
                name TEXT NOT NULL, category TEXT NOT NULL DEFAULT 'Autre',
                acquisition_value REAL NOT NULL DEFAULT 0,
                current_value REAL NOT NULL DEFAULT 0,
                acquisition_date TEXT, notes TEXT DEFAULT '', workspace TEXT DEFAULT 'perso',
                created_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE)''')
        conn.commit()
    except Exception:
        conn.rollback()

    # ── Migration : table asset_catalog ───────────────────────────────────────
    try:
        if is_postgres():
            c.execute('''CREATE TABLE IF NOT EXISTS asset_catalog (
                ticker TEXT PRIMARY KEY, name TEXT NOT NULL,
                category TEXT, subcategory TEXT, region TEXT,
                description TEXT, currency TEXT DEFAULT 'EUR')''')
        else:
            c.execute('''CREATE TABLE IF NOT EXISTS asset_catalog (
                ticker TEXT PRIMARY KEY, name TEXT NOT NULL,
                category TEXT, subcategory TEXT, region TEXT,
                description TEXT, currency TEXT DEFAULT 'EUR')''')
        conn.commit()
    except Exception:
        conn.rollback()

    # ── Migration scoring v2 : vider profil_investisseur si version dépassée ──
    # Version "scoring_v2" = recalibrage 6 axes (mars 2026)
    SCORING_VERSION = 'scoring_v2'
    try:
        if is_postgres():
            c.execute('''
                CREATE TABLE IF NOT EXISTS app_migrations (
                    key   TEXT PRIMARY KEY,
                    value TEXT NOT NULL
                )
            ''')
        else:
            c.execute('''
                CREATE TABLE IF NOT EXISTS app_migrations (
                    key   TEXT PRIMARY KEY,
                    value TEXT NOT NULL
                )
            ''')
        conn.commit()
        p = placeholder()
        c.execute(f'SELECT value FROM app_migrations WHERE key = {p}', ('scoring_version',))
        row = c.fetchone()
        current_ver = row[0] if row else None
        if current_ver != SCORING_VERSION:
            c.execute('DELETE FROM profil_investisseur')
            if row:
                c.execute(f'UPDATE app_migrations SET value = {p} WHERE key = {p}',
                          (SCORING_VERSION, 'scoring_version'))
            else:
                c.execute(f'INSERT INTO app_migrations (key, value) VALUES ({p}, {p})',
                          ('scoring_version', SCORING_VERSION))
            conn.commit()
            print(f"Migration scoring → {SCORING_VERSION} : table profil_investisseur vidée.")
    except Exception as e:
        conn.rollback()
        print(f"Migration scoring warning: {e}")

    # ── Migration : table envelopes (épargne garantie) ───────────────────────
    try:
        if is_postgres():
            c.execute('''
                CREATE TABLE IF NOT EXISTS envelopes (
                    id             SERIAL PRIMARY KEY,
                    user_id        INTEGER NOT NULL,
                    type           TEXT    NOT NULL,
                    nom            TEXT    NOT NULL,
                    solde          REAL    NOT NULL DEFAULT 0,
                    taux_annuel    REAL    NOT NULL DEFAULT 0,
                    plafond        REAL,
                    date_ouverture TEXT,
                    created_at     TIMESTAMP DEFAULT NOW(),
                    FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                )
            ''')
        else:
            c.execute('''
                CREATE TABLE IF NOT EXISTS envelopes (
                    id             INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id        INTEGER NOT NULL,
                    type           TEXT    NOT NULL,
                    nom            TEXT    NOT NULL,
                    solde          REAL    NOT NULL DEFAULT 0,
                    taux_annuel    REAL    NOT NULL DEFAULT 0,
                    plafond        REAL,
                    date_ouverture TEXT,
                    created_at     TEXT    DEFAULT (datetime('now')),
                    FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
                )
            ''')
        conn.commit()
    except Exception:
        conn.rollback()

    conn.close()
    print("Base de données initialisée.")

def fetchall_as_dict(cursor):
    if is_postgres():
        cols = [desc[0] for desc in cursor.description]
        return [dict(zip(cols, row)) for row in cursor.fetchall()]
    else:
        return cursor.fetchall()

def fetchone_as_dict(cursor):
    if is_postgres():
        cols = [desc[0] for desc in cursor.description]
        row = cursor.fetchone()
        return dict(zip(cols, row)) if row else None
    else:
        return cursor.fetchone()

def create_user(email, password_hash):
    p = placeholder()
    conn = get_db()
    try:
        c = conn.cursor()
        c.execute(f'INSERT INTO users (email, password_hash) VALUES ({p}, {p})', (email, password_hash))
        conn.commit()
        return True
    except Exception:
        conn.rollback()
        return False
    finally:
        conn.close()

def get_user_by_email(email):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(f'SELECT * FROM users WHERE email = {p}', (email,))
    row = fetchone_as_dict(c)
    conn.close()
    return row

def get_user_by_id(user_id):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(f'SELECT * FROM users WHERE id = {p}', (user_id,))
    row = fetchone_as_dict(c)
    conn.close()
    return row

def set_onboarding_completed(user_id):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    val = True if is_postgres() else 1
    c.execute(f'UPDATE users SET onboarding_completed = {p} WHERE id = {p}', (val, user_id))
    conn.commit()
    conn.close()

def add_asset(user_id, ticker, name, asset_type, currency, isin='', envelope='', workspace='perso', taux_fixe=None):
    p = placeholder()
    conn = get_db()
    try:
        c = conn.cursor()
        ticker_val = ticker.upper() if ticker and ticker.strip() else None
        c.execute(
            f'''INSERT INTO assets (user_id, isin, ticker, name, asset_type, currency, envelope, workspace, taux_fixe)
               VALUES ({p}, {p}, {p}, {p}, {p}, {p}, {p}, {p}, {p})''',
            (user_id, isin.upper() if isin else '', ticker_val, name, asset_type,
             currency.upper(), envelope or None, workspace or 'perso', taux_fixe)
        )
        conn.commit()
        return True
    except Exception:
        conn.rollback()
        return False
    finally:
        conn.close()

def get_user_assets(user_id, workspace=None):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    if workspace and workspace != 'all':
        c.execute(f'SELECT * FROM assets WHERE user_id = {p} AND workspace = {p} ORDER BY asset_type, name ASC', (user_id, workspace))
    else:
        c.execute(f'SELECT * FROM assets WHERE user_id = {p} ORDER BY asset_type, name ASC', (user_id,))
    rows = fetchall_as_dict(c)
    conn.close()
    return rows

def get_asset_by_id(asset_id, user_id):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(f'SELECT * FROM assets WHERE id = {p} AND user_id = {p}', (asset_id, user_id))
    row = fetchone_as_dict(c)
    conn.close()
    return row

def delete_asset(asset_id, user_id):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(f'DELETE FROM assets WHERE id = {p} AND user_id = {p}', (asset_id, user_id))
    conn.commit()
    conn.close()

def add_purchase(user_id, asset_id, date, shares, price_per_share, fees=0, notes=''):
    p = placeholder()
    total_cost = round(shares * price_per_share, 4)
    conn = get_db()
    c = conn.cursor()
    c.execute(
        f'''INSERT INTO purchases
           (user_id, asset_id, date, shares, price_per_share, total_cost, fees, notes)
           VALUES ({p}, {p}, {p}, {p}, {p}, {p}, {p}, {p})''',
        (user_id, asset_id, date, shares, price_per_share, total_cost, fees, notes)
    )
    conn.commit()
    conn.close()

def add_purchases_bulk(user_id, rows):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    for r in rows:
        total_cost = round(r['shares'] * r['price_per_share'], 4)
        c.execute(
            f'''INSERT INTO purchases
               (user_id, asset_id, date, shares, price_per_share, total_cost, fees, notes)
               VALUES ({p}, {p}, {p}, {p}, {p}, {p}, {p}, {p})''',
            (user_id, r['asset_id'], r['date'], r['shares'],
             r['price_per_share'], total_cost, r.get('fees', 0), r.get('notes', ''))
        )
    conn.commit()
    conn.close()

def get_purchases_by_asset(asset_id, user_id):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(
        f'''SELECT p.*, a.ticker, a.name, a.currency, a.asset_type
           FROM purchases p
           JOIN assets a ON p.asset_id = a.id
           WHERE p.asset_id = {p} AND p.user_id = {p}
           ORDER BY p.date ASC''',
        (asset_id, user_id)
    )
    rows = fetchall_as_dict(c)
    conn.close()
    return rows

def get_all_purchases(user_id):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(
        f'''SELECT p.*, a.ticker, a.name, a.currency, a.isin, a.asset_type, a.envelope
           FROM purchases p
           JOIN assets a ON p.asset_id = a.id
           WHERE p.user_id = {p}
           ORDER BY p.date DESC''',
        (user_id,)
    )
    rows = fetchall_as_dict(c)
    conn.close()
    return rows

def update_asset_envelope(asset_id, user_id, envelope):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(
        f'UPDATE assets SET envelope = {p} WHERE id = {p} AND user_id = {p}',
        (envelope or None, asset_id, user_id)
    )
    conn.commit()
    conn.close()
    return c.rowcount > 0

def get_purchase_by_id(purchase_id, user_id):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(
        f'''SELECT p.*, a.ticker, a.currency
           FROM purchases p
           JOIN assets a ON p.asset_id = a.id
           WHERE p.id = {p} AND p.user_id = {p}''',
        (purchase_id, user_id)
    )
    row = fetchone_as_dict(c)
    conn.close()
    return row

def update_purchase(purchase_id, user_id, date, shares, price_per_share, fees=0, notes=''):
    p = placeholder()
    total_cost = round(shares * price_per_share, 4)
    conn = get_db()
    c = conn.cursor()
    c.execute(
        f'''UPDATE purchases
           SET date={p}, shares={p}, price_per_share={p}, total_cost={p}, fees={p}, notes={p}
           WHERE id={p} AND user_id={p}''',
        (date, shares, price_per_share, total_cost, fees, notes, purchase_id, user_id)
    )
    conn.commit()
    conn.close()

def delete_purchase(purchase_id, user_id):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(f'DELETE FROM purchases WHERE id = {p} AND user_id = {p}', (purchase_id, user_id))
    conn.commit()
    conn.close()

def add_sale(user_id, asset_id, date, shares, price_per_share, fees=0, notes=''):
    p = placeholder()
    total_proceeds = round(shares * price_per_share, 4)
    conn = get_db()
    c = conn.cursor()
    c.execute(
        f'''INSERT INTO sales
           (user_id, asset_id, date, shares, price_per_share, total_proceeds, fees, notes)
           VALUES ({p}, {p}, {p}, {p}, {p}, {p}, {p}, {p})''',
        (user_id, asset_id, date, shares, price_per_share, total_proceeds, fees, notes)
    )
    conn.commit()
    conn.close()

def get_sales_by_asset(asset_id, user_id):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(
        f'''SELECT s.*, a.ticker, a.name, a.currency
           FROM sales s
           JOIN assets a ON s.asset_id = a.id
           WHERE s.asset_id = {p} AND s.user_id = {p}
           ORDER BY s.date ASC''',
        (asset_id, user_id)
    )
    rows = fetchall_as_dict(c)
    conn.close()
    return rows

def get_all_sales(user_id):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(
        f'''SELECT s.*, a.ticker, a.name, a.currency, a.asset_type
           FROM sales s
           JOIN assets a ON s.asset_id = a.id
           WHERE s.user_id = {p}
           ORDER BY s.date DESC''',
        (user_id,)
    )
    rows = fetchall_as_dict(c)
    conn.close()
    return rows

def get_sale_by_id(sale_id, user_id):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(f'SELECT * FROM sales WHERE id = {p} AND user_id = {p}', (sale_id, user_id))
    row = fetchone_as_dict(c)
    conn.close()
    return row

def delete_sale(sale_id, user_id):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(f'DELETE FROM sales WHERE id = {p} AND user_id = {p}', (sale_id, user_id))
    conn.commit()
    conn.close()

def set_dca_goal(user_id, monthly_target):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    if is_postgres():
        c.execute(
            f'''INSERT INTO dca_goals (user_id, monthly_target)
               VALUES ({p}, {p})
               ON CONFLICT(user_id) DO UPDATE SET
                   monthly_target = EXCLUDED.monthly_target,
                   updated_at = NOW()''',
            (user_id, monthly_target)
        )
    else:
        c.execute(
            f'''INSERT INTO dca_goals (user_id, monthly_target)
               VALUES ({p}, {p})
               ON CONFLICT(user_id) DO UPDATE SET
                   monthly_target = excluded.monthly_target,
                   updated_at = datetime('now')''',
            (user_id, monthly_target)
        )
    conn.commit()
    conn.close()

def get_dca_goal(user_id):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(f'SELECT * FROM dca_goals WHERE user_id = {p}', (user_id,))
    row = fetchone_as_dict(c)
    conn.close()
    return row

def add_dividend(user_id, asset_id, date, amount, notes=''):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(
        f'''INSERT INTO dividends (user_id, asset_id, date, amount, notes)
           VALUES ({p}, {p}, {p}, {p}, {p})''',
        (user_id, asset_id, date, round(amount, 4), notes)
    )
    conn.commit()
    conn.close()

def get_dividends_by_asset(asset_id, user_id):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(
        f'''SELECT d.*, a.ticker, a.name, a.currency
           FROM dividends d
           JOIN assets a ON d.asset_id = a.id
           WHERE d.asset_id = {p} AND d.user_id = {p}
           ORDER BY d.date ASC''',
        (asset_id, user_id)
    )
    rows = fetchall_as_dict(c)
    conn.close()
    return rows

def get_all_dividends(user_id):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(
        f'''SELECT d.*, a.ticker, a.name, a.currency, a.asset_type
           FROM dividends d
           JOIN assets a ON d.asset_id = a.id
           WHERE d.user_id = {p}
           ORDER BY d.date DESC''',
        (user_id,)
    )
    rows = fetchall_as_dict(c)
    conn.close()
    return rows

def delete_dividend(dividend_id, user_id):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(f'DELETE FROM dividends WHERE id = {p} AND user_id = {p}', (dividend_id, user_id))
    conn.commit()
    conn.close()

def get_total_dividends(user_id):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(f'SELECT COALESCE(SUM(amount), 0) as total FROM dividends WHERE user_id = {p}', (user_id,))
    row = fetchone_as_dict(c)
    conn.close()
    return round(row['total'], 2)

def update_user_email(user_id, new_email):
    p = placeholder()
    conn = get_db()
    try:
        c = conn.cursor()
        c.execute(f'UPDATE users SET email = {p} WHERE id = {p}', (new_email, user_id))
        conn.commit()
        return True
    except Exception:
        conn.rollback()
        return False
    finally:
        conn.close()

def update_user_password(user_id, new_password_hash):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(f'UPDATE users SET password_hash = {p} WHERE id = {p}', (new_password_hash, user_id))
    conn.commit()
    conn.close()

def create_email_token(user_id, token):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(f'DELETE FROM email_tokens WHERE user_id = {p}', (user_id,))
    c.execute(f'INSERT INTO email_tokens (user_id, token) VALUES ({p}, {p})', (user_id, token))
    conn.commit()
    conn.close()

def get_email_token(token):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(f'SELECT * FROM email_tokens WHERE token = {p}', (token,))
    row = fetchone_as_dict(c)
    conn.close()
    return row

def verify_user_email(user_id):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(f'UPDATE users SET email_verified = TRUE WHERE id = {p}', (user_id,))
    c.execute(f'DELETE FROM email_tokens WHERE user_id = {p}', (user_id,))
    conn.commit()
    conn.close()

def save_profil_investisseur(user_id, score_global, nom_profil, scores_axes, recommandation):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(
        f'''INSERT INTO profil_investisseur (user_id, score_global, nom_profil, scores_axes, recommandation)
           VALUES ({p}, {p}, {p}, {p}, {p})''',
        (user_id, score_global, nom_profil, json.dumps(scores_axes), recommandation)
    )
    conn.commit()
    conn.close()

def get_last_profil_investisseur(user_id):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(
        f'SELECT * FROM profil_investisseur WHERE user_id = {p} ORDER BY id DESC LIMIT 1',
        (user_id,)
    )
    row = fetchone_as_dict(c)
    conn.close()
    if row:
        row = dict(row)
        try:
            row['scores_axes'] = json.loads(row['scores_axes'])
        except Exception:
            row['scores_axes'] = {}
    return row

def get_cached_risk_score(ticker):
    """Returns cached risk score dict if exists and < 24h, else None."""
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(f'SELECT * FROM asset_risk_scores WHERE ticker = {p}', (ticker,))
    row = fetchone_as_dict(c)
    conn.close()
    if not row:
        return None
    row = dict(row)
    date_calc = row.get('date_calcul')
    if date_calc:
        try:
            if isinstance(date_calc, str):
                dt = datetime.datetime.fromisoformat(
                    date_calc.replace('Z', '').split('.')[0].replace('T', ' ')
                )
            else:
                dt = date_calc
                if hasattr(dt, 'tzinfo') and dt.tzinfo:
                    dt = dt.replace(tzinfo=None)
            if (datetime.datetime.utcnow() - dt).total_seconds() >= 86400:
                return None
        except Exception:
            pass
    return {
        'score':      int(row.get('score', 40)),
        'volatilite': row.get('volatilite'),
        'drawdown':   row.get('drawdown'),
        'beta':       row.get('beta'),
        'sharpe':     row.get('sharpe'),
        'var_95':     row.get('var_95'),
        'source':     row.get('source', 'default'),
    }

def save_risk_score(ticker, data):
    """Upsert risk score for a ticker."""
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    score  = data.get('score', 40)
    vol    = data.get('volatilite')
    dd     = data.get('drawdown')
    beta   = data.get('beta')
    sharpe = data.get('sharpe')
    var_95 = data.get('var_95')
    src    = data.get('source', 'default')
    if is_postgres():
        c.execute(f'''
            INSERT INTO asset_risk_scores
                (ticker, score, volatilite, drawdown, beta, sharpe, var_95, date_calcul, source)
            VALUES ({p}, {p}, {p}, {p}, {p}, {p}, {p}, NOW(), {p})
            ON CONFLICT(ticker) DO UPDATE SET
                score={p}, volatilite={p}, drawdown={p}, beta={p},
                sharpe={p}, var_95={p}, date_calcul=NOW(), source={p}
        ''', (ticker, score, vol, dd, beta, sharpe, var_95, src,
              score, vol, dd, beta, sharpe, var_95, src))
    else:
        c.execute(f'''
            INSERT INTO asset_risk_scores
                (ticker, score, volatilite, drawdown, beta, sharpe, var_95, date_calcul, source)
            VALUES ({p}, {p}, {p}, {p}, {p}, {p}, {p}, datetime('now'), {p})
            ON CONFLICT(ticker) DO UPDATE SET
                score=excluded.score, volatilite=excluded.volatilite,
                drawdown=excluded.drawdown, beta=excluded.beta,
                sharpe=excluded.sharpe, var_95=excluded.var_95,
                date_calcul=datetime('now'), source=excluded.source
        ''', (ticker, score, vol, dd, beta, sharpe, var_95, src))
    conn.commit()
    conn.close()

def export_purchases_csv(user_id):
    purchases = get_all_purchases(user_id)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Date', 'Ticker', 'Nom', 'Type', 'ISIN', 'Parts',
                     'Prix/part', 'Total investi', 'Frais', 'Notes'])
    for p in purchases:
        writer.writerow([
            p['date'], p['ticker'], p['name'], p['asset_type'],
            p['isin'], p['shares'], p['price_per_share'],
            p['total_cost'], p['fees'] or 0, p['notes'] or ''
        ])
    return output.getvalue()

def generate_csv_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io as _io

    wb = Workbook()
    ws = wb.active
    ws.title = "Import Achats"

    headers = ['ticker', 'date', 'shares', 'price_per_share', 'fees', 'notes']
    header_fill = PatternFill("solid", fgColor="5B5FED")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    examples = [
        ['ESE.PA',   '17/01/2025', 4, 28.982, 0, 'Achat DCA'],
        ['PAEEM.PA', '04/02/2025', 2, 15.50,  0, 'Achat DCA'],
    ]
    for row_data in examples:
        ws.append(row_data)

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 24

    output = _io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def import_purchases_csv(user_id, file_bytes, filename):
    assets     = get_user_assets(user_id)
    ticker_map = {a['ticker'].upper(): a['id'] for a in assets}
    imported   = 0
    errors     = []
    rows_raw   = []

    if filename.endswith('.xlsx'):
        from openpyxl import load_workbook
        import io as _io
        wb      = load_workbook(filename=_io.BytesIO(file_bytes))
        ws      = wb.active
        headers = [
            str(cell.value).strip().lower() if cell.value is not None else ''
            for cell in ws[1]
        ]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                rows_raw.append(dict(zip(headers, row)))
    else:
        content   = file_bytes.decode('utf-8-sig')
        sample    = content[:1024]
        delimiter = ';' if sample.count(';') > sample.count(',') else ','
        reader    = csv.DictReader(io.StringIO(content), delimiter=delimiter)
        for row in reader:
            rows_raw.append({k.strip().lower(): v for k, v in row.items() if k is not None})

    required = {'ticker', 'date', 'shares', 'price_per_share'}

    for i, row in enumerate(rows_raw, start=2):
        missing = required - set(row.keys())
        if missing:
            errors.append(f'Ligne {i} : colonnes manquantes {missing}')
            continue

        ticker = str(row['ticker']).strip().upper()
        if ticker not in ticker_map:
            errors.append(f'Ligne {i} : ticker "{ticker}" introuvable — ajoute-le d\'abord.')
            continue

        try:
            date_raw = str(row['date']).strip()
            if '/' in date_raw:
                parts = date_raw.split('/')
                date  = f'{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}'
            else:
                date = date_raw[:10]

            def clean_num(val):
                return str(val).replace('€', '').replace(' ', '').replace(',', '.').strip()

            shares          = float(clean_num(row['shares']))
            price_per_share = float(clean_num(row['price_per_share']))
            fees            = float(clean_num(row.get('fees') or 0))
            notes           = str(row.get('notes') or '').strip()

            if shares <= 0 or price_per_share <= 0:
                errors.append(f'Ligne {i} : parts ou prix invalide.')
                continue

            add_purchase(user_id, ticker_map[ticker], date,
                         shares, price_per_share, fees, notes)
            imported += 1

        except (ValueError, KeyError) as e:
            errors.append(f'Ligne {i} : format incorrect ({e})')

    return imported, errors

# ── Reset password tokens ─────────────────────────────────────────────────────

def create_reset_token(user_id, token, expires_at):
    """Store a password reset token, replacing any existing one for this user."""
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(f'DELETE FROM reset_tokens WHERE user_id = {p}', (user_id,))
    expires_str = expires_at if is_postgres() else expires_at.isoformat()
    c.execute(
        f'INSERT INTO reset_tokens (user_id, token, expires_at) VALUES ({p}, {p}, {p})',
        (user_id, token, expires_str)
    )
    conn.commit()
    conn.close()

def get_reset_token(token):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(f'SELECT * FROM reset_tokens WHERE token = {p}', (token,))
    row = fetchone_as_dict(c)
    conn.close()
    return row

def invalidate_reset_token(token):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    used_val = True if is_postgres() else 1
    c.execute(f'UPDATE reset_tokens SET used = {p} WHERE token = {p}', (used_val, token))
    conn.commit()
    conn.close()

# ── Auto dividends cache (Yahoo Finance, TTL 24h) ─────────────────────────────

def get_auto_div_cache(ticker):
    """Returns cached per-share dividend list if < 24h old, else None."""
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(f'SELECT * FROM auto_dividends_cache WHERE ticker = {p}', (ticker,))
    row = fetchone_as_dict(c)
    conn.close()
    if not row:
        return None
    cached_at = row['cached_at']
    try:
        if isinstance(cached_at, str):
            dt = datetime.datetime.fromisoformat(
                cached_at.replace('Z', '').split('.')[0].replace('T', ' ')
            )
        else:
            dt = cached_at
            if hasattr(dt, 'tzinfo') and dt.tzinfo:
                dt = dt.replace(tzinfo=None)
        if (datetime.datetime.utcnow() - dt).total_seconds() >= 86400:
            return None
    except Exception:
        pass
    try:
        return json.loads(row['data_json'])
    except Exception:
        return None

def save_auto_div_cache(ticker, data):
    """Upsert per-share dividend list for a ticker."""
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    data_str = json.dumps(data)
    if is_postgres():
        c.execute(f'''
            INSERT INTO auto_dividends_cache (ticker, data_json)
            VALUES ({p}, {p})
            ON CONFLICT(ticker) DO UPDATE SET data_json=EXCLUDED.data_json, cached_at=NOW()
        ''', (ticker, data_str))
    else:
        c.execute(f'''
            INSERT INTO auto_dividends_cache (ticker, data_json, cached_at)
            VALUES ({p}, {p}, datetime('now'))
            ON CONFLICT(ticker) DO UPDATE SET
                data_json=excluded.data_json,
                cached_at=datetime('now')
        ''', (ticker, data_str))
    conn.commit()
    conn.close()

# ── Alternative Assets CRUD ───────────────────────────────────────────────────

def add_alternative_asset(user_id, name, category, acquisition_value, current_value, acquisition_date, notes='', workspace='perso'):
    p = placeholder()
    conn = get_db()
    try:
        c = conn.cursor()
        c.execute(f'''INSERT INTO alternative_assets (user_id, name, category, acquisition_value, current_value, acquisition_date, notes, workspace)
           VALUES ({p}, {p}, {p}, {p}, {p}, {p}, {p}, {p})''',
            (user_id, name, category, acquisition_value, current_value, acquisition_date, notes, workspace))
        conn.commit()
        return True
    except Exception:
        conn.rollback()
        return False
    finally:
        conn.close()

def get_alternative_assets(user_id, workspace=None):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    if workspace and workspace != 'all':
        c.execute(f'SELECT * FROM alternative_assets WHERE user_id = {p} AND workspace = {p} ORDER BY name ASC', (user_id, workspace))
    else:
        c.execute(f'SELECT * FROM alternative_assets WHERE user_id = {p} ORDER BY name ASC', (user_id,))
    rows = fetchall_as_dict(c)
    conn.close()
    return rows

def update_alternative_asset(asset_id, user_id, name, category, acquisition_value, current_value, acquisition_date, notes='', workspace='perso'):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(f'''UPDATE alternative_assets SET name={p}, category={p}, acquisition_value={p},
       current_value={p}, acquisition_date={p}, notes={p}, workspace={p}
       WHERE id={p} AND user_id={p}''',
        (name, category, acquisition_value, current_value, acquisition_date, notes, workspace, asset_id, user_id))
    conn.commit()
    conn.close()

def delete_alternative_asset(asset_id, user_id):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(f'DELETE FROM alternative_assets WHERE id = {p} AND user_id = {p}', (asset_id, user_id))
    conn.commit()
    conn.close()

# ── Asset Catalog ─────────────────────────────────────────────────────────────

def populate_asset_catalog():
    """Populate asset_catalog with a comprehensive list of tickers. Uses INSERT OR IGNORE / ON CONFLICT DO NOTHING."""
    CATALOG = [
        # ETF MONDE
        ('IWDA.AS','iShares Core MSCI World ETF','ETF','Monde','Mondial','ETF MSCI World All Countries','EUR'),
        ('VWCE.DE','Vanguard FTSE All-World UCITS ETF','ETF','Monde','Mondial','ETF monde entier Vanguard','EUR'),
        ('SWRD.SW','SPDR MSCI World UCITS ETF','ETF','Monde','Mondial','ETF MSCI World SPDR','USD'),
        ('XDWD.DE','Xtrackers MSCI World Swap UCITS ETF','ETF','Monde','Mondial','ETF MSCI World Xtrackers','EUR'),
        ('IUSQ.DE','iShares MSCI ACWI UCITS ETF','ETF','Monde','Mondial','ETF ACWI iShares','EUR'),
        ('HMWO.L','HSBC MSCI World UCITS ETF','ETF','Monde','Mondial','ETF MSCI World HSBC','USD'),
        ('CSPX.AS','iShares Core S&P 500 UCITS ETF','ETF','USA','S&P 500','ETF S&P 500 iShares USD','USD'),
        # ETF USA
        ('SPY','SPDR S&P 500 ETF Trust','ETF','USA','S&P 500','ETF S&P 500 SPDR','USD'),
        ('VOO','Vanguard S&P 500 ETF','ETF','USA','S&P 500','ETF S&P 500 Vanguard','USD'),
        ('QQQ','Invesco QQQ Trust (NASDAQ-100)','ETF','USA','Tech','ETF NASDAQ-100','USD'),
        ('IVV','iShares Core S&P 500 ETF','ETF','USA','S&P 500','ETF S&P 500 iShares','USD'),
        ('VTI','Vanguard Total Stock Market ETF','ETF','USA','Marché Total','ETF marché US total Vanguard','USD'),
        ('ARKK','ARK Innovation ETF','ETF','USA','Innovation','ETF innovation disruptive ARK','USD'),
        ('TQQQ','ProShares UltraPro QQQ','ETF','USA','Levier','ETF NASDAQ-100 x3 levier','USD'),
        ('UPRO','ProShares UltraPro S&P500','ETF','USA','Levier','ETF S&P500 x3 levier','USD'),
        ('SOXL','Direxion Daily Semiconductor Bull 3X','ETF','USA','Levier','ETF semi-conducteurs x3','USD'),
        ('SCHD','Schwab US Dividend Equity ETF','ETF','USA','Dividendes','ETF dividendes US Schwab','USD'),
        # ETF EUROPE
        ('EXW1.DE','iShares Core EURO STOXX 50 UCITS ETF','ETF','Europe','Eurozone','ETF Euro Stoxx 50','EUR'),
        ('DXET.DE','Xtrackers Euro Stoxx 50 UCITS ETF','ETF','Europe','Eurozone','ETF Euro Stoxx 50 Xtrackers','EUR'),
        ('IEUR.AS','iShares Core MSCI Europe UCITS ETF','ETF','Europe','MSCI Europe','ETF Europe iShares','EUR'),
        ('MEUD.PA','Amundi MSCI Europe UCITS ETF','ETF','Europe','MSCI Europe','ETF Europe Amundi','EUR'),
        ('SX5EEX','Euro Stoxx 50 Index','Indice','Europe','Eurozone','Indice Euro Stoxx 50','EUR'),
        # ETF EMERGENTS
        ('EIMI.AS','iShares Core MSCI EM IMI UCITS ETF','ETF','Emergents','Monde émergent','ETF marchés émergents iShares','USD'),
        ('VWO','Vanguard FTSE Emerging Markets ETF','ETF','Emergents','Monde émergent','ETF émergents Vanguard','USD'),
        ('IEMG','iShares Core MSCI Emerging Markets ETF','ETF','Emergents','Monde émergent','ETF émergents iShares','USD'),
        ('PAEEM.PA','Amundi MSCI Emerging Markets UCITS ETF','ETF','Emergents','Monde émergent','ETF émergents Amundi','EUR'),
        ('AEEM.PA','Amundi MSCI Emerging Markets UCITS ETF EUR','ETF','Emergents','Monde émergent','ETF émergents Amundi EUR','EUR'),
        # ETF SECTORIELS
        ('XLK','Technology Select Sector SPDR','ETF','USA','Technologie','ETF secteur tech US','USD'),
        ('XLV','Health Care Select Sector SPDR','ETF','USA','Santé','ETF secteur santé US','USD'),
        ('XLE','Energy Select Sector SPDR','ETF','USA','Énergie','ETF secteur énergie US','USD'),
        ('XLF','Financial Select Sector SPDR','ETF','USA','Finance','ETF secteur financier US','USD'),
        ('XLI','Industrial Select Sector SPDR','ETF','USA','Industrie','ETF secteur industriel US','USD'),
        ('XLP','Consumer Staples Select Sector SPDR','ETF','USA','Conso. de base','ETF consommation de base US','USD'),
        ('VNQ','Vanguard Real Estate ETF','ETF','USA','Immobilier','ETF REIT immobilier US','USD'),
        ('IBB','iShares Biotechnology ETF','ETF','USA','Biotech','ETF biotechnologie','USD'),
        ('SOXX','iShares Semiconductor ETF','ETF','USA','Semi-conducteurs','ETF semi-conducteurs','USD'),
        ('ICLN','iShares Global Clean Energy ETF','ETF','Monde','Énergie propre','ETF énergie propre mondiale','USD'),
        # ETF OBLIGATAIRES
        ('AGGH.L','iShares Core Global Aggregate Bond UCITS ETF','ETF','Monde','Obligations','ETF obligations mondiales agrégées','USD'),
        ('TLT','iShares 20+ Year Treasury Bond ETF','ETF','USA','Obligations','ETF obligations US long terme','USD'),
        ('BND','Vanguard Total Bond Market ETF','ETF','USA','Obligations','ETF obligations US total Vanguard','USD'),
        ('IEAG.AS','iShares Core Euro Aggregate Bond UCITS ETF','ETF','Europe','Obligations','ETF obligations EUR agrégées','EUR'),
        ('EUNR.DE','iShares Euro Govt Bond 15-30yr UCITS ETF','ETF','Europe','Obligations','ETF obligations souveraines EUR LT','EUR'),
        ('HYG','iShares iBoxx High Yield Corporate Bond ETF','ETF','USA','Obligations HY','ETF obligations haut rendement US','USD'),
        ('LQD','iShares iBoxx Investment Grade Corporate Bond ETF','ETF','USA','Obligations IG','ETF obligations IG US','USD'),
        ('VBMFX','Vanguard Total Bond Market Index Fund','ETF','USA','Obligations','Fonds obligations US total Vanguard','USD'),
        # ETF MATIÈRES PREMIÈRES
        ('GLD','SPDR Gold Shares','ETF','Monde','Or','ETF or SPDR','USD'),
        ('SLV','iShares Silver Trust','ETF','Monde','Argent','ETF argent iShares','USD'),
        ('IGLN.L','iShares Physical Gold ETC','ETF','Monde','Or','ETC or physique iShares','USD'),
        ('PDBC','Invesco Optimum Yield Diversified Commodity No K-1','ETF','Monde','Matières premières','ETF matières premières diversifié','USD'),
        ('CPER','United States Copper Index Fund','ETF','Monde','Cuivre','ETF cuivre US','USD'),
        # ETF LEVIERS
        ('LQQ.PA','Amundi Nasdaq-100 Daily (2x) Leveraged UCITS ETF','ETF','USA','Levier','ETF NASDAQ-100 x2 Amundi','EUR'),
        ('CL2.PA','Amundi S&P 500 Daily (2x) Leveraged UCITS ETF','ETF','USA','Levier','ETF S&P500 x2 Amundi','EUR'),
        ('SSO','ProShares Ultra S&P500','ETF','USA','Levier','ETF S&P500 x2 levier','USD'),
        ('QLD','ProShares Ultra QQQ','ETF','USA','Levier','ETF NASDAQ-100 x2 levier','USD'),
        ('SPXL','Direxion Daily S&P 500 Bull 3X Shares','ETF','USA','Levier','ETF S&P500 x3 Direxion','USD'),
        ('BTC3L.L','WisdomTree Bitcoin 3x Daily Leveraged ETP','ETF','Crypto','Levier','ETP Bitcoin x3 WisdomTree','USD'),
        # ETF ESG/ISLAMIQUE
        ('WSRI.L','SPDR Bloomberg SASB Corporate Bond ESG UCITS ETF','ETF','Monde','ESG','ETF ESG obligations SPDR','USD'),
        ('ISDU.L','iShares MSCI World Islamic UCITS ETF','ETF','Monde','Islamique','ETF MSCI World islamique iShares','USD'),
        ('HLAL','Wahed FTSE USA Shariah ETF','ETF','USA','Islamique','ETF Shariah US Wahed','USD'),
        ('ESGW.L','iShares MSCI World ESG Screened UCITS ETF','ETF','Monde','ESG','ETF monde ESG iShares','USD'),
        ('BNPE.PA','BNP Paribas Easy Low Carbon 100 Europe PAB UCITS ETF','ETF','Europe','ESG','ETF ESG Europe BNP Paribas','EUR'),
        ('SUWS.L','UBS MSCI World Socially Responsible UCITS ETF','ETF','Monde','ESG','ETF ISR monde UBS','USD'),
        # ETF GÉOGRAPHIQUES
        ('EWJ','iShares MSCI Japan ETF','ETF','Japon','Japon','ETF Japon iShares','USD'),
        ('INDA','iShares MSCI India ETF','ETF','Inde','Inde','ETF Inde iShares','USD'),
        ('MCHI','iShares MSCI China ETF','ETF','Chine','Chine','ETF Chine iShares','USD'),
        ('EWZ','iShares MSCI Brazil ETF','ETF','Brésil','Brésil','ETF Brésil iShares','USD'),
        ('EZA','iShares MSCI South Africa ETF','ETF','Afrique du Sud','Afrique','ETF Afrique du Sud iShares','USD'),
        ('VGK','Vanguard FTSE Europe ETF','ETF','Europe','Europe','ETF Europe Vanguard','USD'),
        ('EWT','iShares MSCI Taiwan ETF','ETF','Taïwan','Asie','ETF Taïwan iShares','USD'),
        ('EWY','iShares MSCI South Korea ETF','ETF','Corée du Sud','Asie','ETF Corée du Sud iShares','USD'),
        ('ENOR','iShares MSCI Norway ETF','ETF','Norvège','Europe','ETF Norvège iShares','USD'),
        ('EWD','iShares MSCI Sweden ETF','ETF','Suède','Europe','ETF Suède iShares','USD'),
        ('EWQ','iShares MSCI France ETF','ETF','France','Europe','ETF France iShares','USD'),
        ('EWG','iShares MSCI Germany ETF','ETF','Allemagne','Europe','ETF Allemagne iShares','USD'),
        ('EWU','iShares MSCI United Kingdom ETF','ETF','Royaume-Uni','Europe','ETF Royaume-Uni iShares','USD'),
        ('EWA','iShares MSCI Australia ETF','ETF','Australie','Pacifique','ETF Australie iShares','USD'),
        ('EWC','iShares MSCI Canada ETF','ETF','Canada','Amériques','ETF Canada iShares','USD'),
        # ETF DIVIDENDES
        ('VYM','Vanguard High Dividend Yield ETF','ETF','USA','Dividendes','ETF hauts dividendes Vanguard','USD'),
        ('SDY','SPDR S&P Dividend ETF','ETF','USA','Dividendes','ETF dividendes S&P SPDR','USD'),
        ('VHYL.L','Vanguard FTSE All-World High Dividend Yield UCITS ETF','ETF','Monde','Dividendes','ETF dividendes monde Vanguard','USD'),
        ('IDVY.L','iShares Euro Dividend UCITS ETF','ETF','Europe','Dividendes','ETF dividendes Europe iShares','EUR'),
        ('DWX','SPDR S&P International Dividend ETF','ETF','Monde','Dividendes','ETF dividendes international SPDR','USD'),
        ('SPYD','SPDR Portfolio S&P 500 High Dividend ETF','ETF','USA','Dividendes','ETF hauts dividendes S&P 500','USD'),
        # ACTIONS FRANCE (CAC 40 + grandes caps)
        ('MC.PA','LVMH Moët Hennessy Louis Vuitton','Action','France','Luxe','Groupe de luxe français no.1 mondial','EUR'),
        ('TTE.PA','TotalEnergies SE','Action','France','Énergie','Major pétrolier français','EUR'),
        ('AIR.PA','Airbus SE','Action','France','Aéronautique','Leader mondial aéronautique civil','EUR'),
        ('OR.PA',"L'Oréal SA",'Action','France','Cosmétiques','Leader mondial cosmétiques','EUR'),
        ('SAN.PA','Sanofi SA','Action','France','Pharma','Groupe pharmaceutique français','EUR'),
        ('BNP.PA','BNP Paribas SA','Action','France','Banque','Première banque française','EUR'),
        ('AI.PA','Air Liquide SA','Action','France','Chimie','Leader mondial gaz industriels','EUR'),
        ('DG.PA','Vinci SA','Action','France','Infrastructure','Leader mondial concessions infrastructure','EUR'),
        ('RI.PA','Pernod Ricard SA','Action','France','Spiritueux','Groupe de spiritueux mondial','EUR'),
        ('CAP.PA','Capgemini SE','Action','France','Tech','Leader européen services numériques','EUR'),
        ('SU.PA','Schneider Electric SE','Action','France','Énergie','Spécialiste gestion énergie','EUR'),
        ('CS.PA','AXA SA','Action','France','Assurance',"Groupe d'assurance mondial",'EUR'),
        ('BN.PA','Danone SA','Action','France','Agroalimentaire','Groupe agroalimentaire mondial','EUR'),
        ('RMS.PA','Hermès International SA','Action','France','Luxe','Maison de luxe française','EUR'),
        ('KER.PA','Kering SA','Action','France','Luxe','Groupe de luxe (Gucci, YSL)','EUR'),
        ('DSY.PA','Dassault Systèmes SE','Action','France','Logiciels 3D','Logiciels 3D et simulation','EUR'),
        ('SAF.PA','Safran SA','Action','France','Aéronautique','Équipements aéronautiques','EUR'),
        ('MT.PA','ArcelorMittal SA','Action','Europe','Acier','Leader mondial acier','EUR'),
        ('ACA.PA','Crédit Agricole SA','Action','France','Banque','Groupe bancaire français','EUR'),
        ('GLE.PA','Société Générale SA','Action','France','Banque','Banque française internationale','EUR'),
        # ACTIONS US (S&P 500 / grandes caps)
        ('AAPL','Apple Inc.','Action','USA','Tech','Fabricant iPhone, Mac, services','USD'),
        ('MSFT','Microsoft Corporation','Action','USA','Tech','Cloud, Windows, Office, IA','USD'),
        ('GOOGL','Alphabet Inc.','Action','USA','Tech','Google, YouTube, cloud GCP','USD'),
        ('AMZN','Amazon.com Inc.','Action','USA','E-commerce / Cloud','E-commerce et cloud AWS','USD'),
        ('NVDA','NVIDIA Corporation','Action','USA','Semi-conducteurs','Puces GPU, IA, data centers','USD'),
        ('TSLA','Tesla Inc.','Action','USA','Automobile EV','Véhicules électriques et énergie','USD'),
        ('META','Meta Platforms Inc.','Action','USA','Réseaux sociaux','Facebook, Instagram, WhatsApp','USD'),
        ('JPM','JPMorgan Chase & Co.','Action','USA','Banque','Première banque US par actifs','USD'),
        ('V','Visa Inc.','Action','USA','Paiements','Réseau de paiement mondial','USD'),
        ('JNJ','Johnson & Johnson','Action','USA','Pharma / Medtech','Pharma et dispositifs médicaux','USD'),
        ('BRK-B','Berkshire Hathaway Inc.','Action','USA','Conglomérat','Holding de Warren Buffett','USD'),
        ('WMT','Walmart Inc.','Action','USA','Distribution','Plus grand distributeur mondial','USD'),
        ('MA','Mastercard Incorporated','Action','USA','Paiements','Réseau de paiement mondial','USD'),
        ('PG','Procter & Gamble Co.','Action','USA','Conso. de base','Produits de grande consommation','USD'),
        ('HD','The Home Depot Inc.','Action','USA','Distribution','Matériaux de construction US','USD'),
        ('BAC','Bank of America Corp.','Action','USA','Banque','Grande banque américaine','USD'),
        ('XOM','Exxon Mobil Corporation','Action','USA','Énergie','Major pétrolier américain','USD'),
        ('CVX','Chevron Corporation','Action','USA','Énergie','Major pétrolier américain','USD'),
        ('PFE','Pfizer Inc.','Action','USA','Pharma','Groupe pharmaceutique mondial','USD'),
        ('KO','The Coca-Cola Company','Action','USA','Boissons','Leader mondial boissons','USD'),
        ('DIS','The Walt Disney Company','Action','USA','Médias','Médias, parcs, streaming','USD'),
        ('NFLX','Netflix Inc.','Action','USA','Streaming','Plateforme streaming mondiale','USD'),
        ('AMD','Advanced Micro Devices Inc.','Action','USA','Semi-conducteurs','Processeurs CPU et GPU','USD'),
        ('INTC','Intel Corporation','Action','USA','Semi-conducteurs','Fabricant de puces CPU','USD'),
        ('CRM','Salesforce Inc.','Action','USA','SaaS','CRM cloud leader','USD'),
        # ACTIONS EUROPE
        ('ASML.AS','ASML Holding NV','Action','Pays-Bas','Semi-conducteurs','Machines lithographie EUV','EUR'),
        ('NESN.SW','Nestlé SA','Action','Suisse','Agroalimentaire','Plus grand groupe alimentaire mondial','CHF'),
        ('SAP.DE','SAP SE','Action','Allemagne','Logiciels ERP','Leader européen logiciels entreprise','EUR'),
        ('VOW3.DE','Volkswagen AG','Action','Allemagne','Automobile','Groupe automobile mondial VW','EUR'),
        ('BMW.DE','Bayerische Motoren Werke AG','Action','Allemagne','Automobile','Constructeur automobile premium','EUR'),
        ('SIE.DE','Siemens AG','Action','Allemagne','Industrie','Conglomérat industriel allemand','EUR'),
        ('ALV.DE','Allianz SE','Action','Allemagne','Assurance','Premier assureur européen','EUR'),
        # CRYPTO
        ('BTC-USD','Bitcoin','Crypto','Monde','Layer 1','Première cryptomonnaie','USD'),
        ('ETH-USD','Ethereum','Crypto','Monde','Layer 1','Plateforme smart contracts','USD'),
        ('SOL-USD','Solana','Crypto','Monde','Layer 1','Blockchain haute performance','USD'),
        ('BNB-USD','Binance Coin','Crypto','Monde','Exchange','Token natif Binance','USD'),
        ('XRP-USD','XRP (Ripple)','Crypto','Monde','Paiements','Protocole paiements internationaux','USD'),
        ('ADA-USD','Cardano','Crypto','Monde','Layer 1','Blockchain proof-of-stake Cardano','USD'),
        ('AVAX-USD','Avalanche','Crypto','Monde','Layer 1','Blockchain rapide Avalanche','USD'),
        ('DOGE-USD','Dogecoin','Crypto','Monde','Mème coin','Cryptomonnaie mème','USD'),
        ('DOT-USD','Polkadot','Crypto','Monde','Layer 0','Protocole interopérabilité blockchains','USD'),
        ('LINK-USD','Chainlink','Crypto','Monde','Oracle','Oracle blockchain décentralisé','USD'),
        ('MATIC-USD','Polygon','Crypto','Monde','Layer 2','Solution Layer 2 Ethereum','USD'),
        ('LTC-USD','Litecoin','Crypto','Monde','Layer 1','Fork Bitcoin rapide','USD'),
        ('BCH-USD','Bitcoin Cash','Crypto','Monde','Layer 1','Fork Bitcoin','USD'),
        ('ATOM-USD','Cosmos','Crypto','Monde','Layer 0','Écosystème de blockchains interopérables','USD'),
        # ÉPARGNE RÉGLEMENTÉE (sans ticker)
        ('__LIVRET_A','Livret A','Épargne','France','Livret réglementé','Livret réglementé 3.0%/an — sans ticker','EUR'),
        ('__LDDS','LDDS — Livret Développement Durable','Épargne','France','Livret réglementé','Livret réglementé 3.0%/an — sans ticker','EUR'),
        ('__LEP','LEP — Livret Épargne Populaire','Épargne','France','Livret réglementé','Livret réglementé 4.0%/an — sans ticker','EUR'),
        ('__FONDS_EUROS','Fonds euros assurance vie','Épargne','France','Fonds euros','Fonds euros assurance vie ~2.5%/an — sans ticker','EUR'),
        # INDICES
        ('^GSPC','S&P 500 Index','Indice','USA','Grands indices','Indice S&P 500 des 500 grandes caps US','USD'),
        ('^IXIC','NASDAQ Composite','Indice','USA','Grands indices','Indice composite NASDAQ','USD'),
        ('^FCHI','CAC 40 Index','Indice','France','Grands indices','Indice des 40 plus grandes caps françaises','EUR'),
        ('^GDAXI','DAX Performance Index','Indice','Allemagne','Grands indices','Indice des 40 plus grandes caps allemandes','EUR'),
        ('^STOXX50E','Euro Stoxx 50','Indice','Europe','Grands indices','Indice des 50 plus grandes caps eurozone','EUR'),
        ('^N225','Nikkei 225','Indice','Japon','Grands indices','Indice des 225 plus grandes caps japonaises','JPY'),
        ('^HSI','Hang Seng Index','Indice','Hong Kong','Grands indices','Indice de la bourse de Hong Kong','HKD'),
        ('^FTSE','FTSE 100','Indice','Royaume-Uni','Grands indices','Indice des 100 plus grandes caps UK','GBP'),
        ('^RUT','Russell 2000','Indice','USA','Grands indices','Indice des petites caps US','USD'),
    ]

    p = placeholder()
    conn = get_db()
    try:
        c = conn.cursor()
        if is_postgres():
            for row in CATALOG:
                c.execute(f'''INSERT INTO asset_catalog (ticker, name, category, subcategory, region, description, currency)
                    VALUES ({p},{p},{p},{p},{p},{p},{p})
                    ON CONFLICT(ticker) DO NOTHING''', row)
        else:
            for row in CATALOG:
                c.execute(f'''INSERT OR IGNORE INTO asset_catalog (ticker, name, category, subcategory, region, description, currency)
                    VALUES ({p},{p},{p},{p},{p},{p},{p})''', row)
        conn.commit()
        print(f"[catalog] {len(CATALOG)} actifs insérés/mis à jour dans asset_catalog")
    except Exception as e:
        conn.rollback()
        print(f"[catalog] Erreur populate_asset_catalog: {e}")
    finally:
        conn.close()

# ── Enveloppes épargne garantie ───────────────────────────────────────────────

def add_envelope(user_id, type_, nom, solde, taux_annuel, plafond=None, date_ouverture=None):
    p = placeholder()
    conn = get_db()
    try:
        c = conn.cursor()
        c.execute(
            f'''INSERT INTO envelopes (user_id, type, nom, solde, taux_annuel, plafond, date_ouverture)
               VALUES ({p}, {p}, {p}, {p}, {p}, {p}, {p})''',
            (user_id, type_, nom, float(solde), float(taux_annuel),
             float(plafond) if plafond is not None else None, date_ouverture or None)
        )
        conn.commit()
        return True
    except Exception as e:
        conn.rollback()
        print(f"[add_envelope] Erreur: {e}")
        return False
    finally:
        conn.close()

def get_savings_envelopes(user_id):
    p = placeholder()
    conn = get_db()
    c = conn.cursor()
    c.execute(f'SELECT * FROM envelopes WHERE user_id = {p} ORDER BY created_at ASC', (user_id,))
    rows = fetchall_as_dict(c)
    conn.close()
    return [dict(r) for r in rows]

def update_envelope_solde(envelope_id, user_id, solde):
    p = placeholder()
    conn = get_db()
    try:
        c = conn.cursor()
        c.execute(
            f'UPDATE envelopes SET solde = {p} WHERE id = {p} AND user_id = {p}',
            (float(solde), envelope_id, user_id)
        )
        conn.commit()
        return True
    except Exception as e:
        conn.rollback()
        print(f"[update_envelope_solde] Erreur: {e}")
        return False
    finally:
        conn.close()

def delete_envelope(envelope_id, user_id):
    p = placeholder()
    conn = get_db()
    try:
        c = conn.cursor()
        c.execute(f'DELETE FROM envelopes WHERE id = {p} AND user_id = {p}', (envelope_id, user_id))
        conn.commit()
        return True
    except Exception:
        conn.rollback()
        return False
    finally:
        conn.close()
